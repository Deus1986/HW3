import os, zipfile
import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

current_file = os.path.abspath(__file__)
current_dir = os.path.dirname(current_file)
path_to_directory_files = os.path.join(current_dir, "files")
path_to_directory_archive = os.path.join(current_dir, "archive")


@pytest.fixture(scope="function", autouse=False)
def add_files_to_zip():
    if not os.path.exists(path_to_directory_archive):
        os.mkdir(path_to_directory_archive)
    elif not os.path.exists(os.path.join(path_to_directory_archive, "test.zip")):
        file_dir = os.listdir(path_to_directory_files)
        with zipfile.ZipFile(os.path.join(path_to_directory_archive, "test.zip"), mode="w") as zf:
            for file in file_dir:
                add_file = os.path.join(path_to_directory_files, file)
                zf.write(add_file, os.path.basename(add_file))


def test_read_pdf(add_files_to_zip):
    with zipfile.ZipFile(os.path.join(path_to_directory_archive, "test.zip"), "r") as test_zip:
        with test_zip.open("sample-pdf-file.pdf", "r") as pdf:
            reader = PdfReader(pdf).pages
            expected_pdf_text = ("1960s with the release of Letraset sheets containing Lorem Ipsum passages, "
                                 "and more recently with")
            assert expected_pdf_text in reader[0].extract_text()


def test_read_xlsx(add_files_to_zip):
    with zipfile.ZipFile(os.path.join(path_to_directory_archive, "test.zip"), "r") as test_zip:
        with test_zip.open("sample2.xlsx", "r") as xlsx:
            workbook = load_workbook(xlsx).active
            assert workbook.cell(row=2, column=4).value == "Farm subsidies"


def test_read_csv(add_files_to_zip):
    with zipfile.ZipFile(os.path.join(path_to_directory_archive, "test.zip"), "r") as test_zip:
        content = test_zip.read("sample4.csv")
        assert "Game Length" in str(content)
